"""Excel: plantilla descargable, exportar datos e importar desde la misma
plantilla (Plantilla Gastos.xlsx). Cabeceras estables + sinónimos tolerantes."""
import io
import re
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.database import get_db
from app.models.models import Expense, User
from app.routers.auth import _verify_jwt

router = APIRouter(prefix="/excel", tags=["excel"])

HEADERS = ["Fecha", "Descripcion", "Importe (EUR)", "Proposito", "Motivo", "Tipo", "Metodo",
           "Gasto ajeno", "Deudores", "M. devolucion", "Devuelto", "Me corresponde", "Viaje"]

_LISTAS = {
    "A": ("PROPOSITOS", ["Ocio", "Comida", "Bebida", "Transporte", "Estancia", "Ahorro/Inversion",
                        "Productos", "Deporte/Ejercicio", "Farmacia"]),
    "B": ("MOTIVOS", ["Salir", "Planes en casa", "Viajes", "Trabajo", "Estudios", "Evento",
                     "Caprichos", "Mi cumple", "Regalos"]),
    "C": ("TIPOS", ["Recurrente", "Viajes", "Puntual"]),
    "D": ("METODOS", ["Tarjeta", "Bizum", "Split App", "Efectivo", "Deposito", "Online", "Transferencia"]),
}

_META_HEADER = {"fecha": "date", "date": "date", "dia": "date", "fecha (dd/mm/aaaa)": "date",
                "descripcion": "desc", "concepto": "desc", "gasto": "desc", "descripcion del gasto": "desc",
                "importe": "amount", "importe (eur)": "amount", "importe €": "amount", "cantidad": "amount", "eur": "amount",
                "proposito": "purpose", "proposito/uso": "purpose", "categoria": "purpose", "category": "purpose",
                "motivo": "motive", "tipo": "tipo", "tipo de gasto": "tipo",
                "metodo": "method", "metodo pago": "method", "metodo de pago": "method", "pago": "method",
                "gasto ajeno": "ajeno", "deudores": "deudores",
                "m. devolucion": "deuda_metodo", "m. devolución": "deuda_metodo", "metodo devolucion": "deuda_metodo",
                "devuelto": "devuelto", "me corresponde": "me_corresponde", "me toca": "me_corresponde",
                "viaje": "viaje"}
_NORM = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "€": "eur", "/": " "}


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    for a, b in _NORM.items():
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s)


def _header_map(row) -> dict:
    m = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        key = _META_HEADER.get(_norm(str(cell)))
        if key and key not in m:
            m[key] = i
    return m


def _parse_amount(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).strip().replace("€", "").strip()
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        f = float(s)
        return f if f == f else None
    except ValueError:
        return None


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _booly(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = _norm(str(v))
    return s in ("si", "sí", "x", "1", "true", "verdadero")


def _build_workbook(rows: Optional[list[dict]] = None) -> Workbook:
    """Plantilla/exportación: hoja 'Gastos' con cabeceras + 'Datos' con desplegables + 'Resumen'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D9488")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    data = wb.create_sheet("Datos")
    for col, (label, items) in _LISTAS.items():
        data[f"{col}1"] = label
        data[f"{col}1"].font = Font(bold=True)
        for i, it in enumerate(items, start=2):
            data[f"{col}{i}"] = it

    # Desplegables (tipo Propósito / Motivo / Tipo / Método) hasta la fila 1000
    dv_ranges = [("D", "Datos!$A$2:$A$10"), ("E", "Datos!$B$2:$B$10"), ("F", "Datos!$C$2:$C$4"), ("G", "Datos!$D$2:$D$8")]
    for col, formula in dv_ranges:
        dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")

    if rows:
        for r in rows:
            ws.append([
                r.get("date", ""), r.get("desc", ""), r.get("amount", ""), r.get("purpose", ""), r.get("motive", ""),
                r.get("tipo", "") or "Puntual", r.get("method", "") or "Tarjeta",
                "Sí" if r.get("ajeno") else "", r.get("deudores", ""), r.get("deuda_metodo", ""),
                "Sí" if r.get("devuelto") else "", r.get("me_corresponde", ""), r.get("viaje", ""),
            ])

    for idx, w in enumerate([12, 34, 13, 18, 18, 12, 14, 12, 22, 15, 10, 14, 26], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    res = wb.create_sheet("Resumen")
    res.append(["Concepto", "Importe"])
    res["A1"].font = Font(bold=True); res["B1"].font = Font(bold=True)
    hoy = date.today()
    if rows:
        total_mes = sum(r.get("amount", 0) or 0 for r in rows if r.get("date") and str(r["date"])[:7] == str(hoy)[:7])
        total_ano = sum(r.get("amount", 0) or 0 for r in rows if r.get("date") and str(r["date"])[:4] == str(hoy)[:4])
        por_bucket = {}
        for r in rows:
            t = (r.get("tipo") or "").strip().lower()
            if "recurrent" in t:
                b = "Fijo"
            elif "viaj" in t:
                b = "Viajes"
            elif "inversion" in (r.get("purpose") or "").lower() and "transfer" in (r.get("method") or "").lower():
                b = "Inversion"
            else:
                b = "Puntual"
            por_bucket[b] = por_bucket.get(b, 0) + (r.get("amount", 0) or 0)
        res.append(["Total este mes", total_mes])
        res.append(["Total este año", total_ano])
        for b, tot in sorted(por_bucket.items()):
            res.append([f"Total {b}", tot])
    else:
        res.append(["Total este mes", 0])
        res.append(["Total este año", 0])
    return wb


async def _current_user(request: Request, db: AsyncSession) -> User:
    auth = request.headers.get("Authorization", "")
    token = (auth.split(" ", 1)[1] if auth.startswith("Bearer ") else "") or (request.query_params.get("token") or "")
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")
    payload = _verify_jwt(token)
    if not payload or not payload.get("sub"):
        raise HTTPException(status_code=401, detail="Token inválido o expirado")
    user = (await db.execute(select(User).where(User.id == payload["sub"]))).scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return user


@router.get("/plantilla")
async def plantilla(request: Request, db: AsyncSession = Depends(get_db)):
    await _current_user(request, db)
    buf = io.BytesIO()
    _build_workbook().save(buf)
    buf.seek(0)
    return _xlsx_response(buf, "Plantilla Gastos.xlsx")


@router.get("/exportar")
async def exportar(request: Request, db: AsyncSession = Depends(get_db)):
    user = await _current_user(request, db)
    result = await db.execute(
        select(Expense).where(Expense.user_id == user.id).order_by(Expense.date.asc(), Expense.created_at.asc())
    )
    rows = []
    for e in result.scalars().all():
        rows.append({
            "date": e.date, "desc": e.description, "purpose": e.purpose or "", "motive": e.motive or "",
            "tipo": e.tipo or "Puntual", "method": e.method or "Tarjeta", "ajeno": e.ajeno,
            "deudores": e.deudores or "", "deuda_metodo": e.deuda_metodo or "", "devuelto": e.devuelto,
            "me_corresponde": e.me_corresponde, "viaje": e.viaje or "", "amount": e.amount,
        })
    buf = io.BytesIO()
    _build_workbook(rows).save(buf)
    buf.seek(0)
    return _xlsx_response(buf, "Mis gastos.xlsx")


@router.post("/importar")
async def importar(request: Request, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    user = await _current_user(request, db)
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No es un archivo Excel válido (.xlsx)")
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        raise HTTPException(status_code=400, detail="La hoja está vacía")
    m = _header_map(header)
    for req in ("date", "amount", "desc"):
        if req not in m:
            raise HTTPException(status_code=400, detail=f"No encuentro la columna {req} en la cabecera. Revisa la plantilla.")

    existing = set()
    res = await db.execute(select(Expense).where(Expense.user_id == user.id))
    for e in res.scalars().all():
        existing.add((str(e.date), e.description.strip().lower(), round(e.amount, 2)))

    creados = 0
    ignorados = 0
    duplicados = 0
    errores: list[str] = []
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        def gv(key):
            i = m.get(key)
            return row[i] if i is not None and i < len(row) else None
        fecha = _parse_date(gv("date"))
        importe = _parse_amount(gv("amount"))
        desc = (str(gv("desc") or "")).strip()
        if not fecha or importe is None or not desc:
            if any(v is not None for v in row):
                ignorados += 1
            continue
        amount = round(importe, 2)
        if (str(fecha), desc.lower(), amount) in existing:
            duplicados += 1
            continue
        ajeno = _booly(gv("ajeno"))
        devuelto = _booly(gv("devuelto"))
        me = _parse_amount(gv("me_corresponde"))
        if me is None:
            me = amount
        db.add(Expense(
            user_id=user.id, date=fecha, description=desc[:300], amount=amount,
            purpose=(str(gv("purpose") or "")).strip()[:64], motive=(str(gv("motive") or "")).strip()[:64],
            tipo=(str(gv("tipo") or "Puntual")).strip()[:32] or "Puntual",
            method=(str(gv("method") or "Tarjeta")).strip()[:32] or "Tarjeta",
            ajeno=ajeno, deudores=(str(gv("deudores") or "")).strip()[:120],
            deuda_metodo=(str(gv("deuda_metodo") or "")).strip()[:32], devuelto=devuelto,
            me_corresponde=me, viaje=(str(gv("viaje") or "")).strip()[:120],
        ))
        creados += 1
    await db.commit()
    return {"creados": creados, "duplicados": duplicados, "ignoradas": ignorados, "errores": errores[:10]}


def _xlsx_response(buf: io.BytesIO, filename: str):
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
